"""ExcelWriter — streaming output with status-colored rows."""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.utils.dates import now_filename_ts
from app.utils.paths import ensure_dir

if TYPE_CHECKING:
    from pydantic import BaseModel

    from app.core.base_bot import BaseBot

_STATUS_FILL = {
    "ok": "C6EFCE",
    "nao_encontrado": "FFEB9C",
    "erro": "FFCCCC",
    "rate_limit": "FFD580",
    "captcha": "E6E6FA",
    "auth_error": "D9D9D9",
}
_HEADER_FILL = "366092"
_HEADER_FONT_COLOR = "FFFFFF"


class ExcelWriter:
    def __init__(self, output_dir: Path, bot: BaseBot) -> None:
        ensure_dir(output_dir)
        self.bot = bot
        self.columns = bot.output_columns()
        self._center_cols: set[str] = set(bot.center_columns())
        ts = now_filename_ts()
        self.file_path = output_dir / f"resultado_{bot.key}_{ts}.xlsx"
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.title = bot.display_name[:30]
        self._write_header()
        self._wb.save(self.file_path)
        self._rows_written = 0
        logger.info("Planilha de saída criada: {}", self.file_path)

    def _write_header(self) -> None:
        header_font = Font(bold=True, color=_HEADER_FONT_COLOR)
        header_fill = PatternFill(start_color=_HEADER_FILL, end_color=_HEADER_FILL, fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        for idx, name in enumerate(self.columns, 1):
            cell = self._ws.cell(row=1, column=idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            self._ws.column_dimensions[cell.column_letter].width = max(14, min(40, len(name) + 6))
        self._ws.freeze_panes = "A2"

    def append_result(self, result: BaseModel) -> None:
        rows = self.bot.expand_result(result)
        status = getattr(result, "status_consulta", "ok")
        color = _STATUS_FILL.get(str(status), "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        for row_dict in rows:
            row_num = self._ws.max_row + 1
            for col_idx, col_name in enumerate(self.columns, 1):
                value = row_dict.get(col_name, "")
                cell = self._ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill = fill
                if col_name in self._center_cols:
                    cell.alignment = center
            self._rows_written += 1
        self._wb.save(self.file_path)

    @property
    def has_data(self) -> bool:
        return self._rows_written > 0

    def close(self) -> None:
        if self._rows_written == 0:
            try:
                self.file_path.unlink(missing_ok=True)
                logger.info("Nenhuma linha processada — arquivo removido: {}", self.file_path)
            except Exception as exc:
                logger.warning("Falha ao remover arquivo vazio: {}", exc)
            return
        try:
            self._wb.save(self.file_path)
            logger.info("Planilha finalizada: {}", self.file_path)
        except Exception as exc:
            logger.error("Falha ao salvar planilha: {}", exc)
